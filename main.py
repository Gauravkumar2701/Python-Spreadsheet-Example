import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
Product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
product_under_10_inv = {}

# Calculation number of products per supplier

for product_row in range(2, Product_list.max_row + 1):
    supplier_name = Product_list.cell(product_row, 4).value
    inventory = Product_list.cell(product_row, 2).value
    price = Product_list.cell(product_row, 3).value
    product_num = Product_list.cell(product_row, 1).value
    Inventory_total_price = Product_list.cell(product_row, 5)

    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_value = total_value_per_supplier[supplier_name]
        total_value_per_supplier[supplier_name] = current_value + (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Printing product number whose inventory is less then 10
    if inventory < 10:
        product_under_10_inv[int(product_num)] = int(inventory)

    # Adding inventory total price column in the file
    Inventory_total_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(product_under_10_inv)

inv_file.save("file_with_total_price.xlsx")
